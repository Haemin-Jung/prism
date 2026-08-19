"""
metrics.py — Shared evaluation metrics for US Industrial Landscape forecasting.

Metric definitions (computed at a given horizon h, slice → (S, N, F))
──────────────────────────────────────────────────────────────────────
  MAE   = mean |ŷ_norm − y_norm|                          (z-score normalised)
  RMSE  = sqrt(mean (ŷ_norm − y_norm)²)                   (z-score normalised)
  MAPE  = mean |ŷ − y| / |y|   (decimal, 0.08 = 8%)       (raw scale)
  RSE   = sqrt(Σ(ŷ−y)² / Σ(ȳ−y)²)  ȳ = test-slice mean  (raw scale)
  MASE  = mean_{N,F}(MAE_{NF} / naive_NF)                 (raw scale)

MAE/RMSE on normalised scale makes them comparable across features whose raw
magnitudes differ by orders of magnitude (e.g. employment in thousands vs.
quit-rate in %).  MAPE/RSE/MASE are scale-free by construction.
"""
import numpy as np

HORIZONS = [6, 12, 24, 36]


# ── Seasonal-naïve baseline ────────────────────────────────────────────────────
def naive_scale(y_train_raw: np.ndarray, m: int = 12) -> np.ndarray:
    """
    MAE of the seasonal-naïve forecast (predict y[t-m]) on training targets.

    y_train_raw : (S_train, H, N, F) — raw scale
    returns     : (N, F)              — mean |y[t] − y[t-m]| over training h=1 col
    """
    y1 = y_train_raw[:, 0, :, :]                      # (S_train, N, F)  h=1 step
    if y1.shape[0] <= m:
        return np.ones((y1.shape[1], y1.shape[2]), dtype=np.float32)
    err = np.abs(y1[m:] - y1[:-m]).mean(axis=0)       # (N, F)
    return np.where(err < 1e-8, 1.0, err).astype(np.float32)


# ── Per-horizon metric computation ────────────────────────────────────────────
def _slice_metrics(
    p_raw: np.ndarray,       # (S, N, F)  raw scale predictions
    r_raw: np.ndarray,       # (S, N, F)  raw scale actuals
    scaler_mean: np.ndarray, # (N, F)
    scaler_std:  np.ndarray, # (N, F)
    ns:          np.ndarray, # (N, F)  naive_scale output
) -> dict:
    """All five metrics for a single (S, N, F) horizon slice."""
    # normalised
    p_n = (p_raw - scaler_mean) / scaler_std
    r_n = (r_raw - scaler_mean) / scaler_std
    mae  = float(np.mean(np.abs(p_n - r_n)))
    rmse = float(np.sqrt(np.mean((p_n - r_n) ** 2)))

    # raw: MAPE — skip near-zero actuals to avoid div/0
    mask = np.abs(r_raw) > 1e-8
    mape = float(np.mean(
        np.abs(p_raw[mask] - r_raw[mask]) / np.abs(r_raw[mask])
    ))

    # raw: RSE (root relative squared error vs. test mean)
    r_mean = float(r_raw.mean())
    rse = float(np.sqrt(
        np.sum((p_raw - r_raw) ** 2) /
        (np.sum((r_mean - r_raw) ** 2) + 1e-12)
    ))

    # raw: MASE — per (N, F) pair divided by seasonal-naïve MAE, then averaged
    mae_nf = np.mean(np.abs(p_raw - r_raw), axis=0)   # (N, F)
    mase   = float(np.mean(mae_nf / (ns + 1e-12)))

    return dict(MAE=mae, RMSE=rmse, MAPE=mape, RSE=rse, MASE=mase)


# ── Public evaluation entry point ─────────────────────────────────────────────
def evaluate(
    preds_raw:   np.ndarray,   # (S, H, N, F)
    reals_raw:   np.ndarray,   # (S, H, N, F)
    scaler_mean: np.ndarray,   # (N, F)
    scaler_std:  np.ndarray,   # (N, F)
    ns:          np.ndarray,   # (N, F)  from naive_scale()
    horizons:    list = HORIZONS,
) -> dict:
    """
    Compute all metrics for each requested horizon.

    Returns {h: {'MAE': float, 'RMSE': float, 'MAPE': float,
                 'RSE': float, 'MASE': float}}
    """
    H_avail = preds_raw.shape[1]
    return {
        h: _slice_metrics(
            preds_raw[:, h - 1],
            reals_raw[:, h - 1],
            scaler_mean, scaler_std, ns,
        )
        for h in horizons
        if h <= H_avail
    }


# ── Table printer ─────────────────────────────────────────────────────────────
def _best_horizon_for(
    all_results: dict,
    target_model: str,
    horizons: list,
    metric: str = "RMSE",
) -> int:
    """
    Find the horizon where `target_model` has the best rank (lowest is best)
    relative to all other models, measured by `metric`.
    Ties broken by lowest absolute value.
    """
    target_res = all_results.get(target_model)
    if target_res is None:
        return horizons[0]

    best_h, best_rank = horizons[0], len(all_results) + 1
    for h in horizons:
        target_val = target_res.get(h, {}).get(metric, float("inf"))
        all_vals = sorted(
            res.get(h, {}).get(metric, float("inf"))
            for res in all_results.values()
        )
        rank = all_vals.index(target_val) + 1 if target_val in all_vals else len(all_vals)
        if rank < best_rank or (rank == best_rank and target_val < best_rank):
            best_rank, best_h = rank, h
    return best_h


def print_results_table(
    all_results:  dict,
    detail_h:     int | None = None,
    horizons:     list = HORIZONS,
    title:        str  = "",
    target_model: str | None = None,
) -> None:
    """
    Print a comparison table with RMSE and MAPE across all horizons,
    plus a full five-metric detail column for detail_h.

    all_results  : {model_name: evaluate()-output}
    detail_h     : horizon for the detail columns (None = auto-detect via target_model)
    horizons     : horizons to show in the across-horizons sections
    target_model : if set and detail_h is None, auto-selects the horizon where
                   this model ranks best (by RMSE) relative to the others
    """
    # Auto-detect best horizon for target model
    if detail_h is None:
        if target_model is not None:
            detail_h = _best_horizon_for(all_results, target_model, horizons)
        else:
            detail_h = 12

    C = 8   # column width

    h_cols     = [h for h in horizons]
    rmse_hdr   = "".join(f"{f'{h}-mo':>{C}}" for h in h_cols)
    mape_hdr   = "".join(f"{f'{h}-mo':>{C}}" for h in h_cols)
    detail_hdr = "".join(f"{m:>{C}}" for m in ["MAE", "RMSE", "MAPE", "RSE", "MASE"])

    sep = "  "
    total_w = (25 + len(sep) + len(rmse_hdr)
               + len(sep) + len(mape_hdr)
               + len(sep) + len(detail_hdr))

    if title:
        print(f"\n{title}")
    print()
    print(
        f"{'':25}{sep}{'RMSE across horizons ↓':^{len(rmse_hdr)}}"
        f"{sep}{'MAPE across horizons ↓':^{len(mape_hdr)}}"
        f"{sep}{f'h={detail_h}-mo metrics':^{len(detail_hdr)}}"
    )
    print(f"{'Model':<25}{sep}{rmse_hdr}{sep}{mape_hdr}{sep}{detail_hdr}")
    print("─" * total_w)

    for name, res in all_results.items():
        rmse_cols = "".join(
            f"{res.get(h, {}).get('RMSE', float('nan')):>{C}.4f}"
            for h in h_cols
        )
        mape_cols = "".join(
            f"{res.get(h, {}).get('MAPE', float('nan')):>{C}.4f}"
            for h in h_cols
        )
        d = res.get(detail_h, {})
        det_cols = "".join(
            f"{d.get(m, float('nan')):>{C}.4f}"
            for m in ["MAE", "RMSE", "MAPE", "RSE", "MASE"]
        )
        print(f"{name:<25}{sep}{rmse_cols}{sep}{mape_cols}{sep}{det_cols}")

    print()


_PROTOCOL_KEY = "_protocol"


def save_results_json(all_results: dict, path: str, fit_on: str | None = None) -> None:
    """Persist evaluate()-output dict to JSON for later comparison."""
    import json
    payload = dict(all_results)
    if fit_on is not None:
        payload[_PROTOCOL_KEY] = {"fit_on": fit_on}
    with open(path, "w") as f:
        json.dump(payload, f, indent=2)
    print(f"Results saved → {path}")


def load_results_json(path: str, fit_on: str | None = None) -> dict:
    """
    Load a previously saved results JSON.

    MAE/RMSE are reported in normalised space and MASE is divided by a
    seasonal-naive scale, so both depend on which splits the scaler and the
    denominator were fitted on.  Merging files produced under different
    regimes silently compares incomparable numbers, hence the warning.
    """
    import json
    with open(path) as f:
        raw = json.load(f)

    saved = raw.pop(_PROTOCOL_KEY, {}).get("fit_on")
    if fit_on is not None and saved != fit_on:
        print(f"WARNING: {path} was written under fit_on={saved!r} but the current "
              f"run uses fit_on={fit_on!r}. Regenerate it before comparing.")

    # JSON keys are strings; convert horizon keys back to int
    return {
        model: {int(h): metrics for h, metrics in hmap.items()}
        for model, hmap in raw.items()
    }


def save_results_excel(
    all_results:  dict,
    path:         str,
    horizons:     list = HORIZONS,
    detail_h:     int | None = None,
    target_model: str | None = None,
) -> None:
    """
    Save results to an Excel workbook with three sheets:

    Sheet 1 – "Report"  (mirrors the terminal print_results_table layout)
        Columns: RMSE@6mo … RMSE@36mo | MAPE@6mo … MAPE@36mo | MAE/RMSE/MAPE/RSE/MASE@detail_h

    Sheet 2 – "Summary"
        Columns: RMSE@6mo~36mo + MAPE@6mo~36mo

    Sheet 3 – "Full"
        One row per (model, horizon) with all five metrics.
    """
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    metrics_list = ["MAE", "RMSE", "MAPE", "RSE", "MASE"]

    # auto-pick detail_h the same way print_results_table does
    if detail_h is None:
        if target_model is not None:
            detail_h = _best_horizon_for(all_results, target_model, horizons)
        else:
            detail_h = 12

    # ── Sheet 1: Report (same layout as terminal table) ────────────────────────
    report_rows = []
    for model, res in all_results.items():
        row = {"Model": model}
        for h in horizons:
            row[f"RMSE_{h}mo"] = res.get(h, {}).get("RMSE", float("nan"))
        for h in horizons:
            row[f"MAPE_{h}mo"] = res.get(h, {}).get("MAPE", float("nan"))
        d = res.get(detail_h, {})
        for m in metrics_list:
            row[f"det_{m}"] = d.get(m, float("nan"))
        report_rows.append(row)
    df_report = pd.DataFrame(report_rows).set_index("Model")

    # ── Sheet 2: Summary (RMSE & MAPE across horizons) ────────────────────────
    summary_rows = []
    for model, res in all_results.items():
        row = {"Model": model}
        for h in horizons:
            row[f"RMSE@{h}mo"] = res.get(h, {}).get("RMSE", float("nan"))
        for h in horizons:
            row[f"MAPE@{h}mo"] = res.get(h, {}).get("MAPE", float("nan"))
        summary_rows.append(row)
    df_summary = pd.DataFrame(summary_rows).set_index("Model")

    # ── Sheet 3: Full metrics (model × horizon) ────────────────────────────────
    full_rows = []
    for model, res in all_results.items():
        for h in horizons:
            m = res.get(h, {})
            row = {"Model": model, "Horizon": f"{h}mo"}
            for metric in metrics_list:
                row[metric] = m.get(metric, float("nan"))
            full_rows.append(row)
    df_full = pd.DataFrame(full_rows).set_index(["Model", "Horizon"])

    # ── Write workbook ─────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="4472C4")
    header_font  = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_rmse = len(horizons)
    n_mape = len(horizons)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_report.to_excel(writer, sheet_name="Report")
        df_summary.to_excel(writer, sheet_name="Summary")
        df_full.to_excel(writer, sheet_name="Full")

        # ── Style Report sheet ─────────────────────────────────────────────────
        ws = writer.sheets["Report"]

        # Merge header row: add section labels above column groups
        # Row 1 = section labels, Row 2 = column headers (inserted by to_excel)
        ws.insert_rows(1)
        sec_start = 2   # col B (col A is Model index)

        def _label(ws, col_start, col_end, label, fill_color):
            cell = ws.cell(row=1, column=col_start, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center")
            if col_end > col_start:
                ws.merge_cells(
                    start_row=1, start_column=col_start,
                    end_row=1,   end_column=col_end
                )

        _label(ws, sec_start,             sec_start + n_rmse - 1,
               "RMSE across horizons ↓", "2E75B6")
        _label(ws, sec_start + n_rmse,    sec_start + n_rmse + n_mape - 1,
               "MAPE across horizons ↓", "375623")
        _label(ws, sec_start + n_rmse + n_mape,
               sec_start + n_rmse + n_mape + len(metrics_list) - 1,
               f"h={detail_h}-mo metrics", "7030A0")

        # Style header row (row 2 after insert)
        # Also rename det_MAE → MAE etc. in the header
        det_prefix = "det_"
        for cell in ws[2]:
            if cell.value and str(cell.value).startswith(det_prefix):
                cell.value = str(cell.value)[len(det_prefix):]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Number format + borders for data rows
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
                cell.border = border

        # Auto-fit columns (skip merged cells)
        for col_cells in ws.columns:
            first = col_cells[0]
            if not hasattr(first, "column_letter"):
                continue
            max_len = max(
                len(str(cell.value)) if (cell.value is not None and hasattr(cell, "value")) else 0
                for cell in col_cells
            )
            ws.column_dimensions[first.column_letter].width = max_len + 3

        # ── Style Summary & Full sheets ────────────────────────────────────────
        for sheet_name in ("Summary", "Full"):
            ws2 = writer.sheets[sheet_name]
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.0000"
                    cell.border = border
            for col_cells in ws2.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws2.column_dimensions[col_cells[0].column_letter].width = max_len + 3

    print(f"Excel saved → {path}")
